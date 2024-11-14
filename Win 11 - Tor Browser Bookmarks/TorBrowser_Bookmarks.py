#This python script will extract user bookmarks for the Tor Browser.
#
#
#Copyright(C) 2024 Spyder Forensics LLC (www.spyderforensics.com)
#
#This program is free software: you can redistribute it and/or modify
#it under the terms of the GNU General Public License as published by
#the Free Software Foundation, either version 3 of the License, or
#(at your option) any later version.
#
#This program is distributed in the hope that it will be useful,
#but WITHOUT ANY WARRANTY; without even the implied warranty of
#MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#GNU General Public License for more details.
#
#You can view the GNU General Public License at <https://www.gnu.org/licenses/>.
#
# Version History:
# v beta 2024-10-25
# v 1.0 2024-11-14

import sys
import lz4.block
import json
import argparse
import logging
import re
import sqlite3
from datetime import datetime, timezone
import openpyxl
import base64
import os
from pathlib import Path

logger = None

def setup_logger(filename):
    """Sets up the logging"""
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    file_handler = logging.FileHandler(filename, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S %Z (UTC %z)')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger
    
def is_mozlz4_file(file_path):
    """Check if the file is a valid MOZLZ4 file by reading the header."""
    with open(file_path, 'rb') as f:
        header = f.read(8)  # Read the first 8 bytes
        return header == b'mozLz40\0'

def decompress_mozlz4(input_file):
    """Decompress the MOZLZ4 file and return JSON data."""
    try:
        with open(input_file, 'rb') as f:
            data = f.read()
            print(f"    - Decompressing Backup...")    
            logger.info(f"Decompressing Backup")
        if data[:8] != b'mozLz40\0':
            raise ValueError('Invalid MOZLZ4 header')
            logger.error(f"{input_file}: Invalid MOZLZ4 header")
        else:
            print(f"    - Decompression Successfully Completed")
            logger.info(f"Decompression Successfully Completed")
        decompressed_data = lz4.block.decompress(data[8:])
        return json.loads(decompressed_data)
    except (ValueError, lz4.block.LZ4BlockError, json.JSONDecodeError) as e:
        print(f"Error decompressing file: {e}")
        logger.error(f"{input_file}: Error decompressing file: {e}")
        return None

def extract_base64_icon(bookmark_iconuri, bookmark_title, icons_folder, filename):
    """Extract base64-encoded icons and save them to a folder."""
    if bookmark_iconuri.startswith('data:image'):
        file_extension = 'ico' if 'image/x-icon' in bookmark_iconuri else 'png'
        base64_data = bookmark_iconuri.split(',')[1]
        try:
            # Decode the Base64 data
            image_data = base64.b64decode(base64_data)
            # Sanitize the title to create a valid filename
            sanitized_title = re.sub(r'[\ufeff<>:"/\\|?*]', '', bookmark_title)
            sanitized_title = re.sub(r'[:]', '_', sanitized_title)  
            # Create a subfolder based on the filename
            subfolder_path = os.path.join(icons_folder, filename) 
            create_output_directory(subfolder_path)
            output_filename = f"{sanitized_title.replace(' ', '_')}_icon.{file_extension}"  
            icon_output = os.path.join(subfolder_path, output_filename)  
            # Save the image data to the file
            with open(icon_output, 'wb') as image_file:
                image_file.write(image_data)
            logger.info(f"Favicon Extracted - {output_filename}")
            return output_filename, icon_output
        except (base64.binascii.Error, OSError) as e:
            print(f"Error decoding Base64 data for {bookmark_title}: {e}")
            logger.error(f"Error decoding Base64 data for {bookmark_title}: {e}")
            return None 
    return None  

def convert_unix_timestamp(microseconds):
    """Convert Unix timestamp in microseconds to a human-readable date."""
    seconds = microseconds / 1_000_000
    return datetime.fromtimestamp(seconds, timezone.utc).strftime('%Y-%m-%d %H:%M:%S')

def extract_root_info(json_data, folders_found_count):
    """Extract root information from the JSON data."""
    rootfolderid = json_data.get('id')  
    rootfolder = json_data.get('root')
    folder_type = 'folder'  # Returning for excel output
    rootcreated = convert_unix_timestamp(json_data.get('dateAdded', None))
    rootlastmod = convert_unix_timestamp(json_data.get('lastModified', None))
    folders_found_count += 1  # Increment the folder count
    return rootfolderid, rootfolder, rootcreated, rootlastmod, folders_found_count

def extract_folder_info(folder, parent_folder_name=None):
    """Extract folder information from the folder JSON object, including the parent folder."""
    folder_id = folder.get('id')
    folder_type = 'Folder'
    folder_name = folder.get('title')
    folder_description = ''
    folder_url = ''
    folder_added = convert_unix_timestamp(folder.get('dateAdded'))
    folder_modified = convert_unix_timestamp(folder.get('lastModified'))
    parent_name = parent_folder_name or 'PlacesRoot'
    return folder_id, folder_type, parent_name, folder_name, folder_description, folder_url, folder_added, folder_modified

def extract_separator_info(separator, current_folder_name, filename):
    """Extract separator information from the separator JSON object."""
    separator_id = separator.get('id')
    separator_type = 'Visual Separator'
    folder_name = current_folder_name
    separator_title = separator.get('title')
    separator_description = ''
    separator_url = ''
    separator_added = convert_unix_timestamp(separator.get('dateAdded'))
    separator_modified = convert_unix_timestamp(separator.get('lastModified'))
    return separator_id, separator_type, folder_name, separator_title, separator_description, separator_url, separator_added, separator_modified

def extract_json_data(json_data, icons_folder, filename):
    """Extract bookmark information from the JSON data."""
    print(f"    - Extracting Information from JSON File")
    logger.info(f"Extracting Information from JSON File")
    folder_info = []
    bookmark_info = []
    separator_info = []
    favicon_info = []
    summary_info = []
    favicons_found_count = 0
    folders_found_count = 0
    bookmarks_found_count = 0
    separators_found_count = 0

    rootfolderid, rootfolder, rootcreated, rootlastmod, folders_found_count = extract_root_info(json_data, folders_found_count)
    folder_info.append((rootfolderid, 'Folder', '', rootfolder, '', '', rootcreated, rootlastmod))

    def extract_from_folder(folder, parent_folder_name=None):
        nonlocal folders_found_count, bookmarks_found_count, favicons_found_count, separators_found_count

        folder_id, folder_type, parent_name, folder_name, folder_description, folder_url, folder_added, folder_modified = extract_folder_info(
            folder, parent_folder_name
        )
        
        folders_found_count += 1
        folder_info.append((folder_id, folder_type, parent_name, folder_name, folder_description, folder_url, folder_added, folder_modified))

        current_folder_name = folder_name or parent_folder_name or 'Unnamed Folder'

        for child in folder.get('children', []):
            child_type = child.get('type')
            if child_type == 'text/x-moz-place':
                bookmark_data = extract_bookmark_info(child, current_folder_name, filename)
                bookmark_info.append(bookmark_data)
                bookmarks_found_count += 1
                
                favicon_data = extract_favicon_info(child, current_folder_name, icons_folder, filename)
                if favicon_data:
                    favicon_info.append(favicon_data[:-1])
                    favicons_found_count += 1
                    
            elif child_type == 'text/x-moz-place-separator':
                separator_data = extract_separator_info(child, current_folder_name, filename)
                separator_info.append(separator_data)
                separators_found_count += 1

            elif child_type == 'text/x-moz-place-container':
                extract_from_folder(child, current_folder_name)

    for folder in json_data.get('children', []):
        extract_from_folder(folder)

    print(f"        - {folders_found_count} Bookmark Folder(s) Found")
    logger.info(f"{folders_found_count} Bookmark Folder(s) Found")
    print(f"        - {separators_found_count} Visual Separator(s) Found")
    logger.info(f"{separators_found_count} Visual Separator(s) Found")
    print(f"        - {bookmarks_found_count} Bookmark(s) Found")
    logger.info(f"{bookmarks_found_count} Bookmark(s) Found")
    print(f"        - {favicons_found_count} Bookmark Favicon(s) Found")
    logger.info(f"{favicons_found_count} Bookmark Favicon(s) Found")
    summary_info.append((folders_found_count, separators_found_count, bookmarks_found_count, favicons_found_count))
    logger.info(f"Finished Extracting Information from JSON File")
    if favicons_found_count > 0:
        print(f"    - Bookmark Favicons Exported to: {icons_folder}")
        logger.info(f"Bookmark Favicons Exported to: {icons_folder}")
    return summary_info, folder_info, bookmark_info, separator_info, favicon_info

def extract_bookmark_info(bookmark, folder_name, filename):
    """Extract bookmark information and check for favicon."""
    bookmark_id = bookmark.get('id')
    bookmark_type = 'Bookmark'
    bookmark_title = bookmark.get('title', 'Unnamed Bookmark')
    bookmark_uri = bookmark.get('uri', '')
    bookmark_description = bookmark.get('description', '')
    bookmark_iconuri = bookmark.get('iconuri', '')
    bookmark_added = convert_unix_timestamp(bookmark.get('dateAdded', 'unknown'))
    bookmark_modified = convert_unix_timestamp(bookmark.get('lastModified', 'unknown'))
    return (
        bookmark_id, bookmark_type, folder_name, bookmark_title, 
        bookmark_uri, bookmark_description, bookmark_added, 
        bookmark_modified
    )

def extract_favicon_info(favicon, folder_name, icons_folder, filename):
    """Extract favicon information from the bookmark JSON object."""
    favicon_id = favicon.get('id')
    favicon_title = favicon.get('title', 'Unnamed Bookmark')
    favicon_uri = favicon.get('uri', '')
    favicon_iconuri = favicon.get('iconuri', '')

    icon_filename, icon_path = extract_base64_icon(favicon_iconuri, favicon_title, icons_folder, filename) or (None, None)
    has_favicon = icon_path is not None

    if has_favicon:
        return (
            favicon_id, favicon_title, favicon_uri, icon_filename, icon_path, has_favicon
        )
    return None

def create_output_directory(output_folder):
    """Create the output directory if it doesn't exist."""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    elif not os.path.isdir(output_folder):
        logger.error(f"{output_folder} exists but is not a directory")
        raise NotADirectoryError(f"{output_folder} exists but is not a directory")

def extract_backup_date_from_filename(filename):
    """Extract the date from the Bookmarkbackup filename."""
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    print(f"    - Backup Date Extracted from filename")
    logger.info(f"Backup Date Extracted from filename")
    return date_match.group(0) if date_match else 'Unknown'

def write_info_to_sheet(summary_info, folder_info, bookmark_info, separator_info, favicon_info, bookmarks_sheet, favicon_sheet, filename, extractionsummary_sheet):
    """Write extracted information to the Excel sheets"""
    logger.info(f"Writing Extracted Information to Spreadsheet")
    def sanitize_data(data):
        """Ensures data contains only valid Excel-compatible values"""
        return [str(item) if isinstance(item, (tuple, list)) else item for item in data]
    combined_info = []
    if folder_info:
        for folder in folder_info:
            combined_info.append([filename, *folder])
    if bookmark_info:
        for bookmark in bookmark_info:
            combined_info.append([filename, *bookmark])
    if separator_info:
        for separator in separator_info:
            combined_info.append([filename, *separator])
    combined_info.sort(key=lambda x: x[1] if len(x) > 1 else "")
    if combined_info:
        for row in combined_info:
            bookmarks_sheet.append(row)
    if summary_info:
        for summary in summary_info:
            extractionsummary_sheet.append([filename, *summary])
    if favicon_info:
        for favicon in favicon_info:
            sanitized_favicon = sanitize_data(favicon)
            favicon_sheet.append([filename, *sanitized_favicon])

def save_decompressed_json(decompressed_folder, filename, json_data):
    """Save the decompressed JSON file."""
    decompressed_backups_folder = os.path.join(decompressed_folder, "Decompressed Backups")
    create_output_directory(decompressed_backups_folder) 
    output_json_path = os.path.join(decompressed_backups_folder, filename.replace('.jsonlz4', '.json'))
    with open(output_json_path, 'w') as json_file:
        json.dump(json_data, json_file, indent=4)
    print(f"    - Decompressed JSON File Exported to: {output_json_path}")
    logger.info(f"Decompressed JSON File Exported to: {output_json_path}")

def process_single_backup(file_path, filename, decompressed_folder, favicon_sheet, bookmarks_sheet, backup_summary_sheet, icons_folder, missing_recordids, deleted_sheet, extractionsummary_sheet, sqlite_bookmarks):
    """Process a single MOZLZ4 backup file."""
    print(f"\n- Bookmark Backup Name: {filename}")
    logger.info(f"Processing: {filename}")
        
    if is_mozlz4_file(file_path):
        json_data = decompress_mozlz4(file_path)
        if json_data:
            summary_info, folder_info, bookmark_info, separator_info, favicon_info = extract_json_data(json_data, icons_folder, filename)
            backup_date = extract_backup_date_from_filename(filename)            
            write_info_to_sheet(summary_info, folder_info, bookmark_info, separator_info, favicon_info, bookmarks_sheet, favicon_sheet, filename, extractionsummary_sheet)       
            backup_summary_sheet.append([filename, backup_date])
            save_decompressed_json(decompressed_folder, filename, json_data)
            find_missing_records_in_data(filename, folder_info, bookmark_info, separator_info, missing_recordids, deleted_sheet)
            find_overwritten_records_in_data(filename, folder_info, bookmark_info, separator_info, deleted_sheet, sqlite_bookmarks)
            logger.info(f"Finished Processing: {filename}")
        else:
            print(f"- Failed to process JSON data for {filename}")
            logger.error(f"Failed to Process JSON data for {filename}")
    else:
        print(f"- Invalid MOZLZ4 file: {filename}")
        logger.error(f"Invalid MOZLZ4 file: {filename}")

def find_missing_records_in_data(filename, folder_info, bookmark_info, separator_info, missing_recordids, deleted_sheet):
    """Find IDs in the bookmark backups that are missing from places.sqlite or have different URLs."""
    print(f'    - Attempting to Find Missing IDs in the Backup...')
    logger.info(f"Attempting to Find Missing IDs in the Backup")
    folder_ids = {folder[0]: (folder, folder[1]) for folder in folder_info}
    bookmark_ids = {bookmark[0]: (bookmark, bookmark[1]) for bookmark in bookmark_info}
    separator_ids = {separator[0]: (separator, separator[1]) for separator in separator_info}
    combined_ids = {**folder_ids, **bookmark_ids, **separator_ids}
    missing_entries = [id_ for id_ in missing_recordids if id_ in combined_ids]
    if missing_entries:
        for id_ in missing_entries:
            original_data, entry_type = combined_ids[id_]
            print(f"    	- Missing Entry Found: ID: {id_}, Type: {entry_type}, Title: {original_data[3]}")
            logger.info(f"Missing Entry Found: ID: {id_}, Type: {entry_type}, Title: {original_data[3]}")
            deleted_sheet.append([filename, *original_data])
    else:
        print("    	- No Missing IDs Identified")
        logger.info(f"No Missing IDs Identified")
        
def find_overwritten_records_in_data(filename, folder_info, bookmark_info, separator_info, deleted_sheet, sqlite_bookmarks):
    """Identify entries (bookmarks, folders, separators) with the same ID but different type, title, or URL in places.sqlite and the backup."""
    print(f'    - Attempting to Find Reused IDs in the Backup...')
    logger.info(f"Attempting to Find Reused IDs in the Backup")
    backup_entries = {}
    for entry in bookmark_info:
        backup_entries[entry[0]] = {
            "type": entry[1],
            "parent_folder": entry[2],
            "title": entry[3],
            "description": entry[5],
            "url": entry[4],
            "added_date": entry[6],
            "modified_date": entry[7]
        }       
    for entry in folder_info:
        backup_entries[entry[0]] = {
            "type": entry[1],
            "parent_folder": entry[2],
            "title": entry[3],
            "description": None, # Folders don't have descriptions
            "url": None,  # Folders don't have URLs
            "added_date": entry[6],
            "modified_date": entry[7]
        }
    for entry in separator_info:
        backup_entries[entry[0]] = {
            "type": entry[1],
            "parent_folder": entry[2],
            "title": entry[3],
            "description": None, # Separators don't have descriptions
            "url": None,  # Separators don't have URLs
            "added_date": entry[6],
            "modified_date": entry[7]
        } 
    # Flag to track if any reused IDs with differences are found
    reused_ids_found = False    
    # Compare each entry in the backup with the corresponding entry in sqlite_bookmarks
    for entry_id, backup_data in backup_entries.items():
        sqlite_data = sqlite_bookmarks.get(entry_id)
        if sqlite_data:
            changes = []
            # Compare only type, and url fields
            if sqlite_data["type"] != backup_data["type"]:
                changes.append(f"Backup Type: {backup_data['type']}, Backup Title: {backup_data['title']} -> Current Type: {sqlite_data['type']}, Current Title: {sqlite_data.get('title', 'N/A')}")
            if sqlite_data.get("url") != backup_data["url"]:
                changes.append(f"Backup Type: {backup_data['type']}, Backup Title: {backup_data['title']} -> Current Type: {sqlite_data['type']}, Current Title: {sqlite_data.get('title', 'N/A')}")
            # Log changes if any discrepancies were found
            if changes:
                reused_ids_found = True  # Set the flag to True if any changes are found
                print(f"        - ID {entry_id} has been reused: {', '.join(changes)}")
                logger.info(f"ID {entry_id} has been reused: {', '.join(changes)}")
                deleted_sheet.append([
                    filename,                     
                    entry_id,                     
                    backup_data["type"],          
                    backup_data["parent_folder"], 
                    backup_data["title"],         
                    backup_data["url"] if backup_data["url"] else "N/A",  
                    backup_data["description"],   
                    backup_data["added_date"],    
                    backup_data["modified_date"], 
                ])
    if not reused_ids_found:
        print(f"        - No Reused IDs Identified")  
        logger.info(f"No Reused IDs Identified")
        
def process_bookmark_backups(bookmarks_backup_folder, output_folder, bookmarks_sheet, favicon_sheet, backup_summary_sheet, icons_folder, missing_recordids, deleted_sheet, extractionsummary_sheet, sqlite_bookmarks):
    """Process all MOZLZ4 bookmark backups and export to JSON and Excel."""
    create_output_directory(output_folder)

    for filename in os.listdir(bookmarks_backup_folder):
        if filename.endswith('.jsonlz4'):
            file_path = os.path.join(bookmarks_backup_folder, filename)
            process_single_backup(file_path, filename, output_folder, favicon_sheet, bookmarks_sheet, backup_summary_sheet, icons_folder, missing_recordids, deleted_sheet, extractionsummary_sheet, sqlite_bookmarks)

def process_sqlite_database(places_path, favicon_path, bookmarks_sheet, favicon_sheet, icons_folder, placesfilename, faviconsfilename, active_sheet, extractionsummary_sheet):
    """Process the SQLite databases and write results to Excel sheets."""
    try:
        with sqlite3.connect(f"file:{places_path}?mode=ro", uri=True) as conn:
            cursor = conn.cursor()
            cursor.execute("PRAGMA query_only = 1;") #Extra safe guard to ensure no writes to database
            print(f"- Opening places.sqlite (Read-Only)")
            logger.info(f"Opening places.sqlite (Read-Only)")
            bookmark_query = """
            SELECT 
                child.id,
                CASE child.type
                    WHEN 1 THEN 'Bookmark'
                    WHEN 2 THEN 'Folder'
                    WHEN 3 THEN 'Visual Separator'
                    ELSE child.type
                END AS 'Type',
                CASE
                    WHEN child.parent = 1 THEN 'PlacesRoot'
                    ELSE parent.title
                END AS parent_folder,
                CASE 
                    WHEN child.id = 1 THEN 'PlacesRoot'
                    ELSE child.title
                END AS Title,
                moz_places.url, 
                moz_places.description, 
                DATETIME((child.dateAdded / 1000000), 'unixepoch') AS 'Date Added (UTC)',
                DATETIME((child.lastModified / 1000000), 'unixepoch') AS 'Date Last Modified (UTC)'
            FROM moz_bookmarks AS child
            LEFT JOIN moz_places ON child.fk = moz_places.id  
            LEFT JOIN moz_bookmarks AS parent ON child.parent = parent.id;
            """
            cursor.execute(bookmark_query)
            logger.info(f"Executing Query to Extract Bookmark Information")
            bookmark_rows = cursor.fetchall()
            logger.info(f"Writing Extracted Information to Spreadsheet")
            for row in bookmark_rows:  
                bookmarks_sheet.append([placesfilename, *row])
            for row in bookmark_rows:  
                active_sheet.append([placesfilename, *row])
            type_queries = {
                "Bookmark Folder(s)": 2,
                "Visual Separator(s)": 3,
                "Bookmark(s)": 1
            }
            folders_found_count = 0
            separators_found_count = 0
            bookmarks_found_count = 0
            for label, type_value in type_queries.items():
                cursor.execute("SELECT id FROM moz_bookmarks WHERE type = ?", (type_value,))
                rows = cursor.fetchall()
                count = len(rows)
                print(f"    - {count} {label} Found")
                logger.info(f"{count} {label} Found")
                if label == "Bookmark Folder(s)":
                    folders_found_count = count
                elif label == "Visual Separator(s)":
                    separators_found_count = count
                elif label == "Bookmark(s)":
                    bookmarks_found_count = count
            extractionsummary_sheet.append([placesfilename, folders_found_count, separators_found_count, bookmarks_found_count, 'N/A'])
            if Path(favicon_path).exists():
                cursor.execute(f"ATTACH DATABASE 'file:{favicon_path}?mode=ro' AS favicons_db;")              
                cursor.execute("PRAGMA favicons_db.query_only = 1;") #Extra safe guard to ensure no writes to database
                print(f"- Attached favicons.sqlite (Read-Only)")
                logger.info(f"Attached favicons.sqlite (Read-Only)")
                favicon_query = """
                SELECT 
                    moz_bookmarks.id,
                    favicons_db.moz_icons.id,
                    moz_bookmarks.title,
                    moz_places.url,
                    favicons_db.moz_icons.data
                FROM moz_bookmarks
                LEFT JOIN moz_places ON moz_bookmarks.fk = moz_places.id
                LEFT JOIN favicons_db.moz_pages_w_icons 
                    ON moz_places.url_hash = favicons_db.moz_pages_w_icons.page_url_hash
                LEFT JOIN favicons_db.moz_icons_to_pages 
                    ON favicons_db.moz_icons_to_pages.page_id = favicons_db.moz_pages_w_icons.id
                LEFT JOIN favicons_db.moz_icons 
                    ON favicons_db.moz_icons.id = favicons_db.moz_icons_to_pages.icon_id
                WHERE favicons_db.moz_icons.data IS NOT NULL;
                """
                cursor.execute(favicon_query)
                logger.info(f"Executing Query to Extract Bookmark Favicons")
                favicon_rows = cursor.fetchall()
                favicons_count = len(favicon_rows)
                print(f"    - {favicons_count} Bookmark Favicon(s) Found")
                logger.info(f"{favicons_count} Bookmark Favicon(s) Found")
                favicons_folder = Path(icons_folder) / 'faviconsdb'
                favicons_folder.mkdir(parents=True, exist_ok=True)
                for index, (bookmark_id, favicon_id, title, url, favicon_data) in enumerate(favicon_rows):
                    sanitized_title = re.sub(r'[<>:"/\\|?*]', '', title) if title else f"favicon_{index + 1}"
                    file_extension = 'ico' if favicon_data[:4] == b'\x00\x00\x01\x00' else 'png'
                    output_filename = f"{sanitized_title}_{favicon_id}.{file_extension}"
                    icon_output_path = favicons_folder / output_filename
                    try:
                        with open(icon_output_path, 'wb') as icon_file:
                            icon_file.write(favicon_data)
                            logger.info(f"Favicon Extracted - {output_filename}")
                    except OSError as e:
                        print(f"Error saving favicon {bookmark_id}: {e}")
                        logger.error(f"Error saving favicon {bookmark_id}: {e}")
                    favicon_sheet.append([faviconsfilename, bookmark_id, title, url, output_filename, str(favicons_folder)])                
                logger.info(f"Writing Extracted Information to Spreadsheet")
                extractionsummary_sheet.append([faviconsfilename, 'N/A', 'N/A', 'N/A', favicons_count])                
                unassociatedfavicon_query = """
                SELECT 
                    favicons_db.moz_icons.id,
                    favicons_db.moz_icons.icon_url,
                    favicons_db.moz_icons.data
                FROM favicons_db.moz_icons
                LEFT JOIN favicons_db.moz_icons_to_pages 
                    ON favicons_db.moz_icons.id = favicons_db.moz_icons_to_pages.icon_id
                LEFT JOIN favicons_db.moz_pages_w_icons 
                    ON favicons_db.moz_icons_to_pages.page_id = favicons_db.moz_pages_w_icons.id
                LEFT JOIN moz_places 
                    ON favicons_db.moz_pages_w_icons.page_url_hash = moz_places.url_hash
                LEFT JOIN moz_bookmarks 
                    ON moz_places.id = moz_bookmarks.fk
                WHERE moz_bookmarks.id IS NULL;
                """
                cursor.execute(unassociatedfavicon_query)
                logger.info(f"Executing Query to Extract Favicons not associated with Bookmarks")
                unassociatedfavicon_rows = cursor.fetchall()
                unassociatedfavicons_count = len(unassociatedfavicon_rows)
                print(f"    - {unassociatedfavicons_count} Unassociated Bookmark Favicon(s) Extracted")
                logger.info(f"{unassociatedfavicons_count} Unassociated Bookmark Favicon(s) Extracted")
                unassociatedfavicons_folder = Path(icons_folder) / 'faviconsdb - unassociated'
                unassociatedfavicons_folder.mkdir(parents=True, exist_ok=True)
                for index, (favicon_id, icon_url, favicon_data) in enumerate(unassociatedfavicon_rows):
                    unassociated_filename = 'unassociatedfavicon'
                    file_extension = (
                        'ico' if favicon_data[:4] == b'\x00\x00\x01\x00' 
                        else 'svg' if favicon_data[:4] == b'\x3c\x73\x76\x67' 
                        else 'png'
                    )
                    output_filename = f"{unassociated_filename}_{favicon_id}.{file_extension}"
                    icon_output_path = unassociatedfavicons_folder / output_filename
                    try:
                        with open(icon_output_path, 'wb') as icon_file:
                            icon_file.write(favicon_data)
                            logger.info(f"Unassociated Favicon Extracted - {output_filename}")
                    except OSError as e:
                        print(f"Error saving favicon {favicon_id}: {e}")
                        logger.error(f"Error saving favicon {favicon_id}: {e}")
                    favicon_sheet.append([f"{faviconsfilename} (Unassociated)", "", "", icon_url, output_filename, str(unassociatedfavicons_folder)])
                logger.info(f"Writing Extracted Information to Spreadsheet")
                extractionsummary_sheet.append([f"{faviconsfilename} (Unassociated)", 'N/A', 'N/A', 'N/A', unassociatedfavicons_count])          
            else:
                print(f"    - Favicons.sqlite not found at: {favicon_path}")
                logger.error(f"Favicons.sqlite not found at: {favicon_path}")
            print("\nPerforming Analysis to Identify Missing IDs in places.sqlite...")
            logger.info(f"Performing Analysis to Identify Missing IDs in places.sqlite")
            cursor.execute("SELECT id FROM moz_bookmarks")
            logger.info(f"Executing Query to extract ID Values from Moz_bookmarks table")
            record_ids = {id_[0] for id_ in cursor.fetchall()}
            if not record_ids:
                print("No Missing IDs Identified")
                logger.info("No Missing IDs Identified")
                return
            min_id, max_id = min(record_ids), max(record_ids)
            complete_set = set(range(min_id, max_id + 1))
            missing_ids = complete_set - record_ids
            if missing_ids:
                print(f"    - Missing IDs: {sorted(missing_ids)}")
                logger.info(f"Missing IDs: {sorted(missing_ids)}")
            else:
                print("\n- No Missing IDs Identified")
                logger.info(f"No Missing IDs Identified")
            #This is used for comparing SQLite data with the Backup Data
            sqlite_bookmarks = {}
            cursor.execute("""
                SELECT moz_bookmarks.id, 
                       CASE moz_bookmarks.type 
                           WHEN 1 THEN 'Bookmark' 
                           WHEN 2 THEN 'Folder' 
                           WHEN 3 THEN 'Visual Separator' 
                           ELSE moz_bookmarks.type 
                       END AS 'Type', 
                       moz_bookmarks.title, 
                       moz_places.url 
                FROM moz_bookmarks 
                LEFT JOIN moz_places ON moz_bookmarks.fk = moz_places.id
            """)
            for row in cursor.fetchall():
                bookmark_id, item_type, title, url = row
                sqlite_bookmarks[bookmark_id] = {
                    "type": item_type,
                    "title": title,
                    "url": url
                }
            return missing_ids, sqlite_bookmarks
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
        logger.error(f"SQLite error: {e}")

def write_excel(workbook):
    """Set up the headers in the Excel workbook."""
    logger.info(f"Creating the excel file to save Bookmark Information")
    extractionsummary_sheet = workbook.active
    extractionsummary_sheet.title = "Extraction Summary"
    backup_summary_sheet = workbook.create_sheet("Bookmark Backup Information")
    active_sheet = workbook.create_sheet("Active Data")
    deleted_sheet = workbook.create_sheet("Deleted Data")
    bookmarks_sheet = workbook.create_sheet("All Bookmark Info")
    favicon_sheet = workbook.create_sheet("Bookmark Favicons")

    extractionsummary_sheet.append([
        'Filename', 'Bookmark Folders Found', 'Visual Separators Found',
        'Bookmarks Found', 'Favicons Found'])
    backup_summary_sheet.append(['Backup Name', 'Backup Date'])
    active_sheet.append([
        'File', 'Record ID', 'Type', 'Parent Folder', 'Title', 
        'URI', 'Description', 'Date Added (UTC)', 'Date Last Modified (UTC)'
    ])
    deleted_sheet.append([
        'File', 'Record ID', 'Type', 'Parent Folder', 'Title', 
        'URI', 'Description', 'Date Added (UTC)', 'Date Last Modified (UTC)'
    ])
    bookmarks_sheet.append([
        'File', 'Record ID', 'Type', 'Parent Folder', 'Title', 
        'URI', 'Description', 'Date Added (UTC)', 'Date Last Modified (UTC)'
    ])
    favicon_sheet.append([
        'File', 'Record ID', 'Title', 'URL', 'Export Filename', 'Export Location'
    ])

    return {
        'Extraction Summary': extractionsummary_sheet,
        'Bookmark Backup Information': backup_summary_sheet,
        'Active Data': active_sheet,
        'Deleted Data': deleted_sheet,
        'All Bookmark Info': bookmarks_sheet,
        'Bookmark Favicons': favicon_sheet
    }

def count_backup_files(bookmarks_backup_folder):
    """Count the number of bookmark backup files in the bookmarkbackups folder."""
    backup_files = [f for f in os.listdir(bookmarks_backup_folder) if os.path.isfile(os.path.join(bookmarks_backup_folder, f))]
    print(f"    - {len(backup_files)} Bookmark Backup File(s) Found")
    logger.info(f"{len(backup_files)} Bookmark Backup File(s) Found")
    return len(backup_files)

def main():
    # Set up argument parsing
    parser = argparse.ArgumentParser(
        description='Extract Bookmark Information from places.sqlite, favicons.sqlite, and Mozlz4 compressed Bookmark Backups for Tor Browser'
    )
    parser.add_argument('-i', '--input', required=True, help='Input Tor Browser profile folder')
    parser.add_argument('-o', '--output', required=True, help='Output folder for decompressed JSON files, Favicons, and Extraction Report')
    args = parser.parse_args()
    print(r"""
   _____                 _             ______                       _          
  / ____|               | |           |  ____|                     (_)         
 | (___  _ __  _   _  __| | ___ _ __  | |__ ___  _ __ ___ _ __  ___ _  ___ ___ 
  \___ \| '_ \| | | |/ _` |/ _ \ '__| |  __/ _ \| '__/ _ \ '_ \/ __| |/ __/ __|
  ____) | |_) | |_| | (_| |  __/ |    | | | (_) | | |  __/ | | \__ \ | (__\__ \
 |_____/| .__/ \__, |\__,_|\___|_|    |_|  \___/|_|  \___|_| |_|___/_|\___|___/
        | |     __/ |                                                          
        |_|    |___/    

Tor Browser Bookmark Extractor
Version: 1.0 Nov, 2024
Author: Spyder Forensics Training
Website: www.spyderforensics.com
Course: Host-Based Dark Web Forensics
""")
    places_path = os.path.join(args.input, 'places.sqlite')
    favicon_path = os.path.join(args.input, 'favicons.sqlite')
    os.makedirs(args.output, exist_ok=True)
    log_file_path = os.path.join(args.output, "TorBookmarksExtraction.log")
    global logger
    logger = setup_logger(log_file_path)
    # Check if places.sqlite exists before proceeding
    if not os.path.exists(places_path):
        print(f"Error: places.sqlite not found at {places_path}. Please ensure the input is the Tor Browser profile folder")
        logger.error(f"places.sqlite not found at {places_path}. Please ensure the input is the Tor Browser profile folder")
        sys.exit(1)
    icons_folder = os.path.join(args.output, 'Bookmark Favicons')
    decompressed_folder = os.path.join(args.output, 'Decompressed Backups')
    logger.info(f"Creating folders in the Output Directory")
    create_output_directory(icons_folder)
    create_output_directory(decompressed_folder)
    workbook = openpyxl.Workbook()
    sheets = write_excel(workbook)
    extractionsummary_sheet = sheets['Extraction Summary']
    missing_recordids = []
    sqlite_bookmarks = {}
    print(f"Extracting Bookmark Information from SQLite Databases...\n")
    logger.info(f"Extracting Bookmark Information from SQLite Databases")
    # Process places.sqlite and favicons.sqlite
    missing_recordids, sqlite_bookmarks = process_sqlite_database(
        places_path=places_path,
        favicon_path=favicon_path,
        active_sheet=sheets['Active Data'],
        bookmarks_sheet=sheets['All Bookmark Info'],
        favicon_sheet=sheets['Bookmark Favicons'],
        icons_folder=icons_folder,
        placesfilename=os.path.basename(places_path),
        faviconsfilename=os.path.basename(favicon_path),
        extractionsummary_sheet=sheets['Extraction Summary']
    )
    logger.info(f"Finished Extracting Bookmark Information from SQLite Databases")
    # Process bookmark backups
    bookmarks_backup_folder = os.path.join(args.input, 'bookmarkbackups')
    print(f"\nExtracting Bookmark Information from BookmarkBackups Folder...\n")
    logger.info(f"Extracting Bookmark Information from BookmarkBackups Folder")
    if os.path.exists(bookmarks_backup_folder):
        print(f"- Analyzing BookmarkBackups Folder at: {bookmarks_backup_folder}")
        logger.info(f"Analyzing BookmarkBackups Folder at: {bookmarks_backup_folder}")
        count_backup_files(bookmarks_backup_folder)
        process_bookmark_backups(
            extractionsummary_sheet=sheets['Extraction Summary'],
            bookmarks_backup_folder=bookmarks_backup_folder,
            output_folder=args.output,
            bookmarks_sheet=sheets['All Bookmark Info'],
            favicon_sheet=sheets['Bookmark Favicons'],
            backup_summary_sheet=sheets['Bookmark Backup Information'],
            icons_folder=icons_folder,
            missing_recordids=missing_recordids,
            deleted_sheet=sheets['Deleted Data'],
            sqlite_bookmarks=sqlite_bookmarks
        )
    else:
        print(f"- No bookmark backups found at: {bookmarks_backup_folder}")
        logger.info(f"No bookmark backups found at: {bookmarks_backup_folder}")
    # Save the workbook
    excel_output_path = os.path.join(args.output, 'TOR_Bookmarks_Extraction.xlsx')
    workbook.save(excel_output_path)
    print(f"\nTor Bookmark Information Saved: {excel_output_path}")
    logger.info(f"Tor Bookmark Information Saved: {excel_output_path}")

if __name__ == "__main__":
    main()
