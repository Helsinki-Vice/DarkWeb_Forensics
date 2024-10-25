#This python script will extract user bookmarks for the Tor Browser.
#
#
#Copyright(C) 2024 Spyder Forensics LLC (www.spyderforensics.com)
#
#This program is free software: you can redistribute it and/or modify
#it under the terms of the GNU General Public License as published by
#the Free Software Foundation, either version 3 of the License, or
#(at your option) any later version.
#
#This program is distributed in the hope that it will be useful,
#but WITHOUT ANY WARRANTY; without even the implied warranty of
#MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#GNU General Public License for more details.
#
#You can view the GNU General Public License at <https://www.gnu.org/licenses/>.
#
# Version History:
# v beta 2024-10-25

import lz4.block
import json
import argparse
import re
import sqlite3
from datetime import datetime, timezone
import openpyxl
import base64
import os
from pathlib import Path

def is_mozlz4_file(file_path):
    """Check if the file is a valid MOZLZ4 file by reading the header."""
    with open(file_path, 'rb') as f:
        header = f.read(8)  # Read the first 8 bytes
        return header == b'mozLz40\0'

def decompress_mozlz4(input_file):
    """Decompress the MOZLZ4 file and return JSON data."""
    try:
        with open(input_file, 'rb') as f:
            data = f.read()
            print(f"    - Decompressing Backup...")    

        if data[:8] != b'mozLz40\0':
            raise ValueError('Invalid MOZLZ4 header')
        else : print(f"    - Decompression Succesfully Completed")
        decompressed_data = lz4.block.decompress(data[8:])
        
        return json.loads(decompressed_data)

    except (ValueError, lz4.block.LZ4BlockError, json.JSONDecodeError) as e:
        print(f"Error decompressing file: {e}")
        return None

def extract_base64_icon(bookmark_iconuri, bookmark_title, icons_folder, filename):
    """Extract base64-encoded icons and save them to a folder."""
    if bookmark_iconuri.startswith('data:image'):
        file_extension = 'ico' if 'image/x-icon' in bookmark_iconuri else 'png'
        base64_data = bookmark_iconuri.split(',')[1]

        try:
            # Decode the Base64 data
            image_data = base64.b64decode(base64_data)

            # Sanitize the title to create a valid filename
            sanitized_title = re.sub(r'[\ufeff<>:"/\\|?*]', '', bookmark_title)  # Remove invalid characters
            sanitized_title = re.sub(r'[:]', '_', sanitized_title)  # Replace colons with underscores
            
            # Create a subfolder based on the filename
            subfolder_path = os.path.join(icons_folder, filename) 
            create_output_directory(subfolder_path)
            
            output_filename = f"{sanitized_title.replace(' ', '_')}_icon.{file_extension}"  # Use sanitized title
            icon_output = os.path.join(subfolder_path, output_filename)  # Create the full output path

            # Save the image data to the file
            with open(icon_output, 'wb') as image_file:
                image_file.write(image_data)           
            return output_filename, icon_output
        except (base64.binascii.Error, OSError) as e:
            print(f"Error decoding Base64 data for {bookmark_title}: {e}")
            return None  # Return None on failure
    return None  # Return None if the icon URI is invalid
    
def convert_unix_timestamp(microseconds):
    """Convert Unix timestamp in microseconds to a human-readable date."""
    seconds = microseconds / 1_000_000
    return datetime.fromtimestamp(seconds, timezone.utc).strftime('%Y-%m-%d %H:%M:%S')

def extract_backup_summary(json_data,log_list, extractionsummary_sheet):
    """Extract backup summary information from the JSON data."""
    install_date = convert_unix_timestamp(json_data.get('dateAdded', None))
    last_modified_date = convert_unix_timestamp(json_data.get('lastModified', None))
    print(f"        - Basic Information Extracted")
    log_message(f"Basic Backup Information Extracted", log_list, extractionsummary_sheet)
    return install_date, last_modified_date

def extract_folder_info(folder, parent_folder_name=None):
    """Extract folder information from the folder JSON object, including the parent folder."""
    folder_id = folder.get('id')
    folder_type = 'Folder' #returning for excel output
    folder_name = folder.get('title')
    folder_description = '' #key doesnt exist for folders, but returning for excel output
    folder_url = '' #key doesnt exist for folders, but returning for excel output
    folder_added = convert_unix_timestamp(folder.get('dateAdded'))
    folder_modified = convert_unix_timestamp(folder.get('lastModified'))

    # If no parent name is provided, use 'root' for top-level folders
    parent_name = parent_folder_name or 'root'
    return folder_id, folder_type, parent_name, folder_name, folder_description, folder_url, folder_added, folder_modified
    
def extract_separator_info(separator, current_folder_name, filename):
    """Extract folder information from the folder JSON object, including the parent folder."""
    separator_id = separator.get('id')
    separator_type = 'Visual Separator' #returning for excel output
    folder_name = current_folder_name
    separator_title = separator.get('title') #Not populated for separators
    separator_description = '' #key doesnt exist for separators, but returning for excel output
    separator_url = '' #key doesnt exist for separators, but returning for excel output
    separator_added = convert_unix_timestamp(separator.get('dateAdded'))
    separator_modified = convert_unix_timestamp(separator.get('lastModified'))
    
    return separator_id, separator_type, folder_name, separator_title, separator_description, separator_url, separator_added, separator_modified

def extract_json_data(json_data, icons_folder, filename, log_list, extractionsummary_sheet):
    """Extract bookmark folders and their bookmarks from the JSON data."""
    print(f"    - Extracting Information from JSON File")
    folder_info = []
    bookmark_info = []
    separator_info = []
    favicon_info = []
    favicons_found_count = 0
    folders_found_count = 0
    bookmarks_found_count = 0
    separators_found_count = 0
    install_date, torlastmod_date = extract_backup_summary(json_data, log_list, extractionsummary_sheet)

    def extract_from_folder(folder, parent_folder_name=None):
        nonlocal folders_found_count, bookmarks_found_count, favicons_found_count, separators_found_count, log_list, extractionsummary_sheet

        folder_id, folder_type, parent_name, folder_name, folder_description, folder_url, folder_added, folder_modified = extract_folder_info(
            folder, parent_folder_name
        )
        
        # Append the folder info
        folder_info.append((folder_id, folder_type, parent_name, folder_name, folder_description, folder_url, folder_added, folder_modified))
        folders_found_count += 1

        # Set the current folder name
        current_folder_name = folder_name or parent_folder_name or 'Unnamed Folder'

        # Iterate over child elements in the folder
        for child in folder.get('children', []):
            child_type = child.get('type')
            if child_type == 'text/x-moz-place':  # Bookmark
                bookmark_data = extract_bookmark_info(child, current_folder_name, filename)
                bookmark_info.append(bookmark_data)
                bookmarks_found_count += 1
                
                # Extract favicon data
                favicon_data = extract_favicon_info(child, current_folder_name, icons_folder, filename)
                if favicon_data:  # Only append if favicon_data is not None
                    favicon_info.append(favicon_data[:-1])
                    favicons_found_count += 1
                    
            elif child_type == 'text/x-moz-place-separator':  # Visual Separator
                separator_data = extract_separator_info(child, current_folder_name, filename)
                separator_info.append(separator_data)
                separators_found_count += 1

            elif child_type == 'text/x-moz-place-container':  # Nested folder
                extract_from_folder(child, current_folder_name)

    # Start extracting from top-level folders
    for folder in json_data.get('children', []):
        extract_from_folder(folder)

    print(f"        - {folders_found_count} Bookmark Folder(s) Found")
    log_message(f"- {folders_found_count} Bookmark Folder(s) Found", log_list, extractionsummary_sheet)
    print(f"        - {separators_found_count} Visual Separator(s) Found")
    log_message(f"- {separators_found_count} Visual Separator(s) Found", log_list, extractionsummary_sheet)
    print(f"        - {bookmarks_found_count} Bookmark(s) Found")
    log_message(f"- {bookmarks_found_count} Bookmark(s) Found", log_list, extractionsummary_sheet)
    print(f"        - {favicons_found_count} Bookmark Favicon(s) Found")
    log_message(f"- {favicons_found_count} Bookmark Favicon(s) Found", log_list, extractionsummary_sheet)
    
    if favicons_found_count > 0:
        print(f"    - Bookmark Favicons Exported to: {icons_folder}")
        log_message(f"- Bookmark Favicons Exported to: {icons_folder}", log_list, extractionsummary_sheet)
    return install_date, torlastmod_date, folder_info, bookmark_info, separator_info, favicon_info

def extract_bookmark_info(bookmark, folder_name, filename):
    """Extract bookmark information and check for favicon."""
    bookmark_id = bookmark.get('id')
    bookmark_type = 'Bookmark'  #Returning for excel output
    bookmark_title = bookmark.get('title', 'Unnamed Bookmark')
    bookmark_uri = bookmark.get('uri', '')
    bookmark_description = bookmark.get('description', '')
    bookmark_iconuri = bookmark.get('iconuri', '')
    bookmark_added = convert_unix_timestamp(bookmark.get('dateAdded', 'unknown'))
    bookmark_modified = convert_unix_timestamp(bookmark.get('lastModified', 'unknown'))

    return (
        bookmark_id, bookmark_type, folder_name, bookmark_title, 
        bookmark_uri, bookmark_description, bookmark_added, 
        bookmark_modified
    )

def extract_favicon_info(favicon, folder_name, icons_folder, filename):
    """Extract bookmark information and check for favicon."""
    favicon_id = favicon.get('id')
    favicon_title = favicon.get('title', 'Unnamed Bookmark')
    favicon_uri = favicon.get('uri', '')
    favicon_iconuri = favicon.get('iconuri', '')

    icon_filename, icon_path = extract_base64_icon(favicon_iconuri, favicon_title, icons_folder, filename) or (None, None)
    has_favicon = icon_path is not None

    # Only return data if a favicon exists
    has_favicon = icon_path is not None
    if has_favicon:
        return (
            favicon_id, favicon_title, favicon_uri, icon_filename, icon_path, has_favicon
        )
    return None  # Return None if no favicon is found

def create_output_directory(output_folder):
    """Create the output directory if it doesn't exist."""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    elif not os.path.isdir(output_folder):
        raise NotADirectoryError(f"{output_folder} exists but is not a directory.")

def extract_backup_date_from_filename(filename):
    """Extract the date from the Bookmarkbackup filename."""
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    print(f"    - Backup Date Extracted from filename")
    return date_match.group(0) if date_match else 'Unknown'

def write_info_to_sheet(folder_info, bookmark_info, separator_info, favicon_info, bookmarks_sheet, favicon_sheet, filename):
    """Write folder, bookmark, separator, and favicon information extracted from backups to the Excel sheets."""
    
    def sanitize_data(data):
        """Ensure data contains only valid Excel-compatible values."""
        return [str(item) if isinstance(item, (tuple, list)) else item for item in data]

    # Prepare a combined list to hold all entries
    combined_info = []

    # Process folder information
    if folder_info:
        for folder in folder_info:
            combined_info.append([filename, *folder])  # Append folder data

    # Process bookmark information
    if bookmark_info:
        for bookmark in bookmark_info:
            combined_info.append([filename, *bookmark])  # Append bookmark data

    # Process separator information
    if separator_info:
        for separator in separator_info:
            combined_info.append([filename, *separator])  # Append separator data
	
    # Sort combined information by ID
    combined_info.sort(key=lambda x: x[1] if len(x) > 1 else "")
    # Write all combined bookmarks information to the bookmarks sheet
    if combined_info:
        for row in combined_info:
            bookmarks_sheet.append(row)
    # Write favicon information to the favicon sheet
    if favicon_info:
        for favicon in favicon_info:
            sanitized_favicon = sanitize_data(favicon)
            favicon_sheet.append([filename, *sanitized_favicon])  # Append favicon data

def save_decompressed_json(decompressed_folder, filename, json_data, log_list, extractionsummary_sheet):
    """Save the decompressed JSON file."""
    decompressed_backups_folder = os.path.join(decompressed_folder, "Decompressed Backups")
    create_output_directory(decompressed_backups_folder) 
    output_json_path = os.path.join(decompressed_backups_folder, filename.replace('.jsonlz4', '.json'))
    with open(output_json_path, 'w') as json_file:
        json.dump(json_data, json_file, indent=4)
    print(f"    - JSON Exported to: {output_json_path}")
    log_message(f"- JSON Exported to: {output_json_path}", log_list, extractionsummary_sheet)

def process_single_backup(file_path, filename, decompressed_folder, favicon_sheet, bookmarks_sheet, backup_summary_sheet, icons_folder, missing_recordids, deleted_sheet, log_list, extractionsummary_sheet):
    """Process a single MOZLZ4 backup file."""
    print(f"\n- Bookmark Backup Name: {filename}")
    log_message(f"Bookmark Backup Name: {filename}", log_list, extractionsummary_sheet)
        
    if is_mozlz4_file(file_path):
        json_data = decompress_mozlz4(file_path)
        if json_data:
            install_date, last_modified_date, folder_info, bookmark_info, separator_info, favicon_info = extract_json_data(json_data, icons_folder, filename, log_list, extractionsummary_sheet)

            # Write folder, separator, and bookmark information to the sheets
            write_info_to_sheet(folder_info, bookmark_info, separator_info, favicon_info, bookmarks_sheet, favicon_sheet, filename)       

            # Add summary information to the backup summary sheet
            backup_date = extract_backup_date_from_filename(filename)
            backup_summary_sheet.append([filename, backup_date, install_date, last_modified_date])
       
            # Save the decompressed JSON
            save_decompressed_json(decompressed_folder, filename, json_data, log_list, extractionsummary_sheet)
            print(f'    - Attempting to Find Missing Entries in the Backup..')
            log_message(f"- Attempting to Find Missing Entries in the Backup", log_list, extractionsummary_sheet)
            missing_entries = find_missing_records_in_data(filename, folder_info, bookmark_info, separator_info, missing_recordids, deleted_sheet, log_list, extractionsummary_sheet)
                
        else:
            print(f"- Failed to process JSON data for {filename}") 
    else:
        print(f"- Invalid MOZLZ4 file: {filename}")

def find_missing_records_in_data(filename, folder_info, bookmark_info, separator_info, missing_recordids, deleted_sheet, log_list, extractionsummary_sheet):
    """Find IDs in the bookmark backups that are missing from the places.sqlite"""
    
    # Extract IDs and their corresponding original data and types from folder_info
    folder_ids = {folder[0]: (folder, folder[1]) for folder in folder_info}  # (original data, type from index 2)

    # Extract IDs and their corresponding original data and types from bookmark_info
    bookmark_ids = {bookmark[0]: (bookmark, bookmark[1]) for bookmark in bookmark_info}  # (original data, type from index 2)

    # Extract IDs and their corresponding original data and types from separator_info
    separator_ids = {separator[0]: (separator, separator[1]) for separator in separator_info}  # (original data, type from index 2)

    # Combine all IDs into a single dictionary for easy lookup
    combined_ids = {**folder_ids, **bookmark_ids, **separator_ids}

    # Find missing record IDs that are present in combined_ids
    missing_entries = [id_ for id_ in missing_recordids if id_ in combined_ids]

    # Prepare to gather details for the missing entries
    missing_details = []

    # Process missing entries if found
    if missing_entries:
        for id_ in missing_entries:
            original_data, entry_type = combined_ids[id_]
            print(f"    	- Missing Entry Found: ID: {id_}, Type: {entry_type}, Title: {original_data[3]}")
            log_message(f"- Missing Entry Found: ID: {id_}, Type: {entry_type}, Title: {original_data[3]}", log_list, extractionsummary_sheet)
           
            # Write all original data to the deleted_sheet
            deleted_sheet.append([filename, *original_data])
    else:
        print("    	- No Missing Entries Found")
        log_message(f"- No Missing Entries Found", log_list, extractionsummary_sheet)

def process_bookmark_backups(bookmarks_backup_folder, output_folder, bookmarks_sheet, favicon_sheet, backup_summary_sheet, icons_folder, missing_recordids, deleted_sheet, log_list, extractionsummary_sheet):
    """Process all MOZLZ4 bookmark backups and export to JSON and Excel."""
    create_output_directory(output_folder)

    for filename in os.listdir(bookmarks_backup_folder):
        if filename.endswith('.jsonlz4'):
            file_path = os.path.join(bookmarks_backup_folder, filename)
            process_single_backup(file_path, filename, output_folder, favicon_sheet, bookmarks_sheet, backup_summary_sheet, icons_folder, missing_recordids, deleted_sheet, log_list, extractionsummary_sheet)

def process_sqlite_database(places_path, favicon_path, bookmarks_sheet, favicon_sheet, icons_folder, placesfilename, faviconsfilename, active_sheet, log_list, extractionsummary_sheet):
    """Process the SQLite databases and write results to Excel sheets."""
    try:
        # Open the places.sqlite database in read-only mode
        with sqlite3.connect(f"file:{places_path}?mode=ro", uri=True) as conn:
            cursor = conn.cursor()
            print(f"- Opening places.sqlite (Read-Only)")
            log_message(f"places.sqlite opened (Read-Only)", log_list, extractionsummary_sheet)

            # Query to extract all bookmark-related data
            bookmark_query = """
            SELECT 
                child.id,
                CASE child.type
                    WHEN 1 THEN 'Bookmark'
                    WHEN 2 THEN 'Folder'
                    WHEN 3 THEN 'Visual Separator'
                    ELSE child.type
                END AS 'Type',
                CASE
                    WHEN child.parent = 1 THEN 'root'
                    ELSE parent.title
                END AS parent_folder,
                CASE 
                    WHEN child.id = 1 THEN 'root'
                    ELSE child.title
                END AS Title,
                moz_places.url, 
                moz_places.description, 
                DATETIME((child.dateAdded / 1000000), 'unixepoch') AS 'Date Added (UTC)',
                DATETIME((child.lastModified / 1000000), 'unixepoch') AS 'Date Last Modified (UTC)'
            FROM moz_bookmarks AS child
            LEFT JOIN moz_places ON child.fk = moz_places.id  
            LEFT JOIN moz_bookmarks AS parent ON child.parent = parent.id;
            """
            cursor.execute(bookmark_query)
            bookmark_rows = cursor.fetchall()

            for row in bookmark_rows:
                bookmarks_sheet.append([placesfilename, *row])  # Properly indented

            # Assuming active_sheet is defined somewhere above
            for row in bookmark_rows:
                active_sheet.append([placesfilename, *row])

            # Query to count different record types
            type_queries = {
                "Bookmark Folder(s)": 2,
                "Visual Separator(s)": 3,
                "Bookmark(s)": 1
            }

            for label, type_value in type_queries.items():
                cursor.execute("SELECT id FROM moz_bookmarks WHERE type = ?", (type_value,))
                rows = cursor.fetchall()
                print(f"    - {len(rows)} {label} Found")
                log_message(f"- {len(rows)} {label} Found", log_list, extractionsummary_sheet)

            # Attach and process favicons.sqlite if it exists
            if Path(favicon_path).exists():

                cursor.execute(f"ATTACH DATABASE 'file:{favicon_path}?mode=ro' AS favicons_db;")
                print(f"- Attached favicons.sqlite (Read-Only)")
                log_message(f"Attached favicons.sqlite (Read-Only)", log_list, extractionsummary_sheet)
                favicon_query = """
                SELECT 
                    moz_bookmarks.id,
                    favicons_db.moz_icons.id,
                    moz_bookmarks.title,
                    moz_places.url,
                    favicons_db.moz_icons.data
                FROM moz_bookmarks
                LEFT JOIN moz_places ON moz_bookmarks.fk = moz_places.id
                LEFT JOIN favicons_db.moz_pages_w_icons 
                    ON moz_places.url_hash = favicons_db.moz_pages_w_icons.page_url_hash
                LEFT JOIN favicons_db.moz_icons_to_pages 
                    ON favicons_db.moz_icons_to_pages.page_id = favicons_db.moz_pages_w_icons.id
                LEFT JOIN favicons_db.moz_icons 
                    ON favicons_db.moz_icons.id = favicons_db.moz_icons_to_pages.icon_id
                WHERE favicons_db.moz_icons.data IS NOT NULL;
                """
                cursor.execute(favicon_query)
                favicon_rows = cursor.fetchall()
                print(f"    - {len(favicon_rows)} Bookmark Favicon(s) Extracted")
                log_message(f"- {len(favicon_rows)} Bookmark Favicon(s) Extracted", log_list, extractionsummary_sheet)

                # Create the favicons output folder
                favicons_folder = Path(icons_folder) / 'faviconsdb'
                favicons_folder.mkdir(parents=True, exist_ok=True)

                for index, (bookmark_id, favicon_id, title, url, favicon_data) in enumerate(favicon_rows):
                    sanitized_title = re.sub(r'[<>:"/\\|?*]', '', title) if title else f"favicon_{index + 1}"
                    file_extension = 'ico' if favicon_data[:4] == b'\x00\x00\x01\x00' else 'png'
                    output_filename = f"{sanitized_title}_{favicon_id}.{file_extension}"
                    icon_output_path = favicons_folder / output_filename

                    try:
                        with open(icon_output_path, 'wb') as icon_file:
                            icon_file.write(favicon_data)
                    except OSError as e:
                        print(f"Error saving favicon {bookmark_id}: {e}")
                        log_message(f" - Error saving favicon {bookmark_id}: {e}", log_list, extractionsummary_sheet)

                    # Write to the favicon sheet
                    favicon_sheet.append([faviconsfilename, bookmark_id, title, url, output_filename, str(favicons_folder)])
            else:
                print(f"    - Favicons.sqlite not found at: {favicon_path}")
                log_message(f"Favicons.sqlite not found at: {favicon_path}", log_list, extractionsummary_sheet)

            # Analyze missing entries in moz_bookmarks table
            print("\nPerforming Analysis to Identify Missing Entries in places.sqlite..")
            log_message("Analysis performed to Identify Missing Entries in places.sqlite", log_list, extractionsummary_sheet)
            cursor.execute("SELECT id FROM moz_bookmarks")
            record_ids = {id_[0] for id_ in cursor.fetchall()}

            if not record_ids:
                print("No IDs found in the moz_bookmarks table.")
                log_message("No IDs found in the moz_bookmarks table.", log_list, extractionsummary_sheet)
                return

            min_id, max_id = min(record_ids), max(record_ids)
            complete_set = set(range(min_id, max_id + 1))
            missing_ids = complete_set - record_ids

            if missing_ids:
                print(f"\n- Missing Entries: {sorted(missing_ids)}")
                log_message(f"- Missing Entries: {sorted(missing_ids)}", log_list, extractionsummary_sheet)
            else:
                print("\n- No Missing Entries Identified.")
                log_message("- No Missing Entries Identified", log_list, extractionsummary_sheet)

            return missing_ids

    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
        log_message(f"SQLite error: {e}", log_list, extractionsummary_sheet)
     
def create_output_directory(directory):
    """Create the output directory if it doesn't exist."""
    os.makedirs(directory, exist_ok=True)

def write_excel(workbook):
    """Set up the headers in the Excel workbook."""
    
    extractionsummary_sheet = workbook.active
    extractionsummary_sheet.title = "Extraction Summary"
    backup_summary_sheet = workbook.create_sheet("Bookmark Backup Information")
    active_sheet = workbook.create_sheet("Active Data")
    deleted_sheet = workbook.create_sheet("Deleted Data")
    bookmarks_sheet = workbook.create_sheet("All Bookmark Info")
    favicon_sheet = workbook.create_sheet("Bookmark Favicons")

    # Headers for each tab
    backup_summary_sheet.append(['Backup Name', 'Backup Date', 'Tor Install Date (UTC)', 'Tor Last Update Date (UTC)'])
    active_sheet.append([
        'File', 'Record ID', 'Type', 'Parent Folder', 'Title', 
        'URI', 'Description', 'Date Added (UTC)', 'Date Last Modified (UTC)'
    ])
    deleted_sheet.append([
        'File', 'Record ID', 'Type', 'Parent Folder', 'Title', 
        'URI', 'Description', 'Date Added (UTC)', 'Date Last Modified (UTC)'
    ])
    bookmarks_sheet.append([
        'File', 'Record ID', 'Type', 'Parent Folder', 'Title', 
        'URI', 'Description', 'Date Added (UTC)', 'Date Last Modified (UTC)'
    ])
    favicon_sheet.append([
        'File', 'Record ID', 'Title', 'URL', 'Export Filename', 'Export Location'
    ])

    return {
        'Extraction Summary': extractionsummary_sheet,
        'Bookmark Backup Information': backup_summary_sheet,
        'Active Data': active_sheet,
        'Deleted Data': deleted_sheet,
        'All Bookmark Info': bookmarks_sheet,
        'Bookmark Favicons': favicon_sheet
    }
    
def count_backup_files(bookmarks_backup_folder, log_list, extractionsummary_sheet):
    """Count the number of bookmark backup files in the bookmarkbackups folder."""
    
    # List all files in the backup folder
    backup_files = [f for f in os.listdir(bookmarks_backup_folder) if os.path.isfile(os.path.join(bookmarks_backup_folder, f))]

    print(f"    - {len(backup_files)} Bookmark Backup File(s) Found")
    log_message(f"{len(backup_files)} Bookmark Backup File(s) Found", log_list, extractionsummary_sheet)
    return len(backup_files)
    
def log_message(message, log_list, extractionsummary_sheet):
    """log messages to the extraction summary sheet"""
    extractionsummary_sheet.append([message])  # Append message to Excel sheet
	
def main():
    # Set up argument parsing
    parser = argparse.ArgumentParser(
        description='Extract Bookmark Informtion from places.sqlite, favicons.sqlite and Mozlz4 compressed Bookmark Backups for Tor Browser'
    )
    parser.add_argument('-i', '--input', required=True, help='Input Tor Browser profile folder')
    parser.add_argument('-o', '--output', required=True, help='Output folder for decompressed JSON files, Favicons, and Extraction Report')
    args = parser.parse_args()

    # Create the output directory and necessary subdirectories
    os.makedirs(args.output, exist_ok=True)
    icons_folder = os.path.join(args.output, 'Bookmark Favicons')
    decompressed_folder = os.path.join(args.output, 'Decompressed Backups')
    create_output_directory(icons_folder)
    create_output_directory(decompressed_folder)
    log_list = []  # Initialize a log list

    # Initialize the Excel workbook and write headers
    workbook = openpyxl.Workbook()
    sheets = write_excel(workbook)
    extractionsummary_sheet = sheets['Extraction Summary']

    print(r"""
   _____                 _             ______                       _          
  / ____|               | |           |  ____|                     (_)         
 | (___  _ __  _   _  __| | ___ _ __  | |__ ___  _ __ ___ _ __  ___ _  ___ ___ 
  \___ \| '_ \| | | |/ _` |/ _ \ '__| |  __/ _ \| '__/ _ \ '_ \/ __| |/ __/ __|
  ____) | |_) | |_| | (_| |  __/ |    | | | (_) | | |  __/ | | \__ \ | (__\__ \
 |_____/| .__/ \__, |\__,_|\___|_|    |_|  \___/|_|  \___|_| |_|___/_|\___|___/
        | |     __/ |                                                          
        |_|    |___/    

Tor Browser Bookmark Extractor
Version: Beta Oct, 2024
Author: Spyder Forensics Training
Website: www.spyderforensics.com
""")

    # Process sqlite databases if they exist
    places_path = os.path.join(args.input, 'places.sqlite')
    favicon_path = os.path.join(args.input, 'favicons.sqlite')
    missing_recordids = []
    print(f"Extracting Bookmark Information from SQLite Databases..\n")
    log_message(f"Bookmark Information from SQLite Databases", log_list, extractionsummary_sheet)

    if os.path.exists(places_path):
        missing_recordids = process_sqlite_database(
            places_path=places_path,
            favicon_path=favicon_path,
            active_sheet=sheets['Active Data'],
            bookmarks_sheet=sheets['All Bookmark Info'],
            favicon_sheet=sheets['Bookmark Favicons'],
            icons_folder=icons_folder,
            placesfilename=os.path.basename(places_path),
            faviconsfilename=os.path.basename(favicon_path),
            log_list=log_list,
            extractionsummary_sheet=sheets['Extraction Summary']
        )
    else:
        print(f"places.sqlite not found at: {places_path}")
        log_message(f"places.sqlite not found at: {places_path}", log_list, extractionsummary_sheet)

    # Process bookmark backups if the folder exists
    bookmarks_backup_folder = os.path.join(args.input, 'bookmarkbackups')
    print(f"\nExtracting Bookmark Information from BookmarkBackups Folder..\n")
    log_message(f"Bookmark Information from BookmarkBackups Folder", log_list, extractionsummary_sheet)

    if os.path.exists(bookmarks_backup_folder):
        print(f"- Analyzing BookmarkBackups Folder at: {bookmarks_backup_folder}")
        log_message(f"BookmarkBackups Folder: {bookmarks_backup_folder}", log_list, extractionsummary_sheet)
        count_backup_files(bookmarks_backup_folder, log_list, extractionsummary_sheet)
        process_bookmark_backups(
            bookmarks_backup_folder=bookmarks_backup_folder,
            output_folder=args.output,
            bookmarks_sheet=sheets['All Bookmark Info'],
            favicon_sheet=sheets['Bookmark Favicons'],
            backup_summary_sheet=sheets['Bookmark Backup Information'],
            icons_folder=icons_folder,
            missing_recordids=missing_recordids,
            deleted_sheet=sheets['Deleted Data'],
            log_list=log_list,
            extractionsummary_sheet=sheets['Extraction Summary']
        )
    else:
        log_message(f"No bookmark backups found at: {bookmarks_backup_folder}", log_list, extractionsummary_sheet)

    # Save the Excel workbook
    excel_output_path = os.path.join(args.output, 'TOR_Bookmarks_Extraction.xlsx')
    workbook.save(excel_output_path)
    print(f"\nTor Bookmark Information Saved: {excel_output_path}")

if __name__ == "__main__":
    main()
